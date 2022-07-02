#pip install openpyxl - in cmd
#install openpyxl in Project- Interpreter
#i p w s r c p p f f p p

import openpyxl
path="F:\Selenium With Python\Data.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
#sheet=workbook.get_sheet_by_name('sheet1')
row=sheet.max_row
col=sheet.max_column
print(row)
print(col)

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value,end="    ")
    print()