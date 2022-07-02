import openpyxl

def GetRow(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_row)

def GetColumn(file,sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_column)

def ReadRow(file,sheet_name,rowno,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell (row=rowno,column=columnno).value

def WriteData(file,sheet_name,rowno,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell (row=rowno,column=columnno).value=data
    workbook.save(file)
