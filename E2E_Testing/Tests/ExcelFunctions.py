import openpyxl

def GetRowCount(File, SheetName):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return (Sheet.max_row)

def GetColumnCount(File, SheetName):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return (Sheet.max_column)

def ReadData(File,SheetName,RowNum,ColumnNum):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    return Sheet.cell(row=RowNum, column=ColumnNum).value

def WriteData(File, SheetName, RowNum, ColumnNum, Data):
    workbook = openpyxl.load_workbook(File)
    Sheet = workbook.get_sheet_by_name(SheetName)
    Sheet.cell(row=RowNum, column=ColumnNum).value = Data
    workbook.save(File)

